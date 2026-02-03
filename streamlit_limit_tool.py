import streamlit as st
import pandas as pd
import openpyxl
import re
from datetime import datetime
from dataclasses import dataclass
from typing import Optional, Tuple, List, Dict

# -----------------------------
# Config
# -----------------------------
WORKBOOK_PATH = "_Broker LOB & Limit Request Logic.xlsx"  # place next to this app, or update path


# -----------------------------
# Utilities
# -----------------------------
_money_re = re.compile(r"[\$,]")

def _to_float_maybe(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip()
        if s == "":
            return None
        s = _money_re.sub("", s)
        # handle M / MM / K
        mult = 1.0
        s_up = s.upper()
        if s_up.endswith(("MM", "M")):
            mult = 1_000_000.0
            s = re.sub(r"(MM|M)$", "", s_up).strip()
        elif s_up.endswith("K"):
            mult = 1_000.0
            s = s_up[:-1].strip()
        try:
            return float(s) * mult
        except Exception:
            return None
    return None

def fmt_limit(x) -> str:
    """Pretty-print limits. Keeps explicit strings like '$1MM/$2MM' intact."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return "—"
    if isinstance(x, str):
        return x.strip()
    if isinstance(x, (int, float)):
        v = float(x)
        if v >= 1_000_000 and v % 1_000_000 == 0:
            return f"${int(v/1_000_000)}M"
        if v >= 1_000:
            return f"${v:,.0f}"
        return str(v)
    return str(x)

@dataclass
class RangeBand:
    label: str
    lo: Optional[float]
    hi: Optional[float]  # inclusive upper bound
    lo_inclusive: bool = True
    hi_inclusive: bool = True

def parse_money_range(label: str) -> RangeBand:
    """
    Parses labels like:
      '$0- $1M', '$1M - $5M', '$5-10M', '$12- $25M', '$25-$35', '$35-$50M', '> $50M'
    Returns RangeBand with lo/hi in dollars.
    """
    s = str(label).strip()
    s = s.replace("—", "-").replace("–", "-")
    s = re.sub(r"\s+", " ", s)

    # Greater-than / Less-than
    m = re.match(r"^>\s*\$?\s*(.+)$", s)
    if m:
        lo = _to_float_maybe(m.group(1))
        return RangeBand(label=s, lo=lo, hi=None, lo_inclusive=False)

    m = re.match(r"^<\s*\$?\s*(.+)$", s)
    if m:
        hi = _to_float_maybe(m.group(1))
        return RangeBand(label=s, lo=None, hi=hi, hi_inclusive=False)

    # Normal ranges
    if "-" in s:
        parts = [p.strip() for p in s.split("-") if p.strip()]
        if len(parts) >= 2:
            lo = _to_float_maybe(parts[0])
            hi = _to_float_maybe(parts[1])
            return RangeBand(label=s, lo=lo, hi=hi)
    # Fallback single value
    v = _to_float_maybe(s)
    return RangeBand(label=s, lo=v, hi=v)

def in_band(x: float, band: RangeBand) -> bool:
    if band.lo is not None:
        if band.lo_inclusive:
            if x < band.lo:
                return False
        else:
            if x <= band.lo:
                return False
    if band.hi is not None:
        if band.hi_inclusive:
            if x > band.hi:
                return False
        else:
            if x >= band.hi:
                return False
    return True

def employee_band(n: int) -> str:
    if n <= 5: return "1-5"
    if n <= 10: return "6-10"
    if n <= 20: return "11-20"
    if n <= 50: return "21-50"
    return ">51"

def state_group(primary_state: str) -> str:
    s = (primary_state or "").strip().upper()
    if s == "CA":
        return "CA"
    if s in {"NY", "IL", "NJ", "MA", "WA"}:
        return "Employee-friendly"
    return "Other"


# -----------------------------
# Data loading
# -----------------------------
@st.cache_data(show_spinner=False)
def load_tables(workbook_path: str):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    def df_sheet(name):
        ws = wb[name]
        data = [list(r) for r in ws.iter_rows(values_only=True)]
        df = pd.DataFrame(data).dropna(how="all").dropna(axis=1, how="all")
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
        return df

    # Core tables
    cyber = df_sheet("Cyber & E&O GTM limit vs. UW li")
    dno = df_sheet("D&O GTM limit vs. UW limit")
    epl_uw = df_sheet("EPL GTM limit vs. UW limit")
    crime = df_sheet("Crime GTM limit")
    gl = df_sheet("GL GTM limit vs. UW limit")

    # EPL recommended limit table (state group x employee band) from exposures sheet
    ws = wb["Exposures & Limit Recommendatio"]

    # These row numbers are based on the current workbook structure.
    # If the sheet changes, search for 'EPL' and re-anchor.
    CA_VALUES_ROW = 115
    EF_VALUES_ROW = 119
    OTHER_VALUES_ROW = 123

    def row_vals(r):
        vals = []
        for c in range(3, 8):  # 5 columns: 1-5, 6-10, 11-20, 21-50, >51
            vals.append(ws.cell(r, c).value)
        return vals

    epl_reco = {
        "CA": dict(zip(["1-5", "6-10", "11-20", "21-50", ">51"], row_vals(CA_VALUES_ROW))),
        "Employee-friendly": dict(zip(["1-5", "6-10", "11-20", "21-50", ">51"], row_vals(EF_VALUES_ROW))),
        "Other": dict(zip(["1-5", "6-10", "11-20", "21-50", ">51"], row_vals(OTHER_VALUES_ROW))),
    }

    return cyber, dno, epl_uw, crime, gl, epl_reco


# -----------------------------
# Recommendation logic
# -----------------------------
def lookup_by_money_band(df: pd.DataFrame, band_col: str, x: float) -> pd.Series:
    bands = [parse_money_range(v) for v in df[band_col].astype(str).tolist()]
    for i, b in enumerate(bands):
        if in_band(x, b):
            return df.iloc[i]
    # If no match, fall back to last row (often the max band)
    return df.iloc[-1]

def lookup_by_employee_band(df: pd.DataFrame, band_col: str, n: int) -> pd.Series:
    # band_col contains strings like '1-25', '26-50', '51-250', '> 1,000'
    def parse_emp_range(s: str) -> Tuple[Optional[int], Optional[int]]:
        s = str(s).strip().replace(",", "")
        if s.startswith(">"):
            lo = int(re.findall(r"\d+", s)[0])
            return lo, None
        if "-" in s:
            a, b = [t.strip() for t in s.split("-")[:2]]
            return int(re.findall(r"\d+", a)[0]), int(re.findall(r"\d+", b)[0])
        nums = re.findall(r"\d+", s)
        if nums:
            v = int(nums[0])
            return v, v
        return None, None

    for i, raw in enumerate(df[band_col].tolist()):
        lo, hi = parse_emp_range(raw)
        if lo is None and hi is None:
            continue
        if hi is None and n > lo:
            return df.iloc[i]
        if hi is not None and lo <= n <= hi:
            return df.iloc[i]
    return df.iloc[-1]

def recommend_gl(gl_df: pd.DataFrame, revenue: float, employees: int, biz_type: str) -> Dict[str, str]:
    row = lookup_by_money_band(gl_df, "Revenue", revenue)
    reco_col = "Recommended Limit (software)" if biz_type != "Hardware / physical products" else "Recommended Limit (hardware)"
    reco = row[reco_col]

    # Offered limit in GL sheet is tied to employee count thresholds; use revenue row + employee count condition.
    # We'll compute offered as the max offered limit among rules whose employee condition matches.
    offered = None
    offers = []
    for _, r in gl_df.iterrows():
        # match same revenue band
        if str(r["Revenue"]) != str(row["Revenue"]):
            continue
        cond = r.get("Employee Count")
        if isinstance(cond, datetime):
            # Excel mis-parsed '10-25' as a date; treat any datetime here as '10-25'
            cond = "10-25"
        cond_s = str(cond).strip() if cond is not None else ""
        ok = True
        if cond_s.startswith("<"):
            m = re.findall(r"\d+", cond_s)
            if m:
                ok = employees < int(m[0])
        elif cond_s.startswith(">"):
            m = re.findall(r"\d+", cond_s)
            if m:
                ok = employees > int(m[0])
        elif cond_s == "10-25":
            ok = 10 <= employees <= 25
        # else: blank => applies generally
        if ok:
            offers.append(r.get("Offered Limit"))
    if offers:
        offered = max([o for o in offers if o is not None], default=None)

    return {"Recommended": fmt_limit(reco), "Offered (UW)": fmt_limit(offered)}

def recommend_cyber(cyber_df: pd.DataFrame, revenue: float) -> Dict[str, str]:
    row = lookup_by_money_band(cyber_df, "Next Twelve months Revenue", revenue)
    return {
        "Recommended (GTM)": fmt_limit(row["Recommended Limit"]),
        "Offered (UW) 1st party": fmt_limit(row["1st party Cyber Offered Limit"]),
        "Offered (UW) 3rd party": fmt_limit(row["3rd Party Cyber offered"]),
    }

def recommend_dno(dno_df: pd.DataFrame, capital_raised: float) -> Dict[str, str]:
    row = lookup_by_money_band(dno_df, "Capital raised", capital_raised)
    return {
        "Recommended (GTM)": fmt_limit(row["Recommended Limit"]),
        "Offered (UW)": fmt_limit(row["Offered Limit"]),
    }

def recommend_epl(epl_uw_df: pd.DataFrame, epl_reco_map: Dict[str, Dict[str, str]], employees: int,
                  primary_state: str, pct_in_ca: Optional[float]) -> Dict[str, str]:
    group = state_group(primary_state)
    # If they have any CA footprint, bias to CA group (more conservative)
    ca_fte = None
    if pct_in_ca is not None and pct_in_ca > 0:
        ca_fte = max(1, int(round(employees * (pct_in_ca / 100.0))))
        group = "CA"

    n_for_band = ca_fte if group == "CA" and ca_fte is not None else employees
    band = employee_band(n_for_band)

    reco = epl_reco_map[group][band]

    uw_row = lookup_by_employee_band(epl_uw_df, "Num Employees", employees)
    return {
        "State group (derived)": group,
        "Employees used for banding": str(n_for_band),
        "Band": band,
        "Recommended (Exposures table)": fmt_limit(reco),
        "Offered (UW)": fmt_limit(uw_row["Offered Limit"]),
    }

def recommend_crime(crime_df: pd.DataFrame, capital_raised: float, regulated: bool) -> Dict[str, str]:
    row = lookup_by_money_band(crime_df, "Capital raised", capital_raised)
    col = "Regulated" if regulated else "Non regulated"
    return {"Recommended (GTM)": fmt_limit(row[col])}


# -----------------------------
# UI
# -----------------------------
st.set_page_config(page_title="Limit Recommendation Tool", layout="wide")

st.title("Limit Recommendation Tool")
st.caption("Interactive prototype driven by the workbook logic (GTM recommended vs. UW offered where available).")

with st.sidebar:
    st.header("Inputs")
    revenue = st.number_input("Next 12 months revenue (USD)", min_value=0.0, value=2_000_000.0, step=100_000.0, format="%.0f")
    employees = st.number_input("Total employees (FTE)", min_value=0, value=20, step=1)

    biz_type = st.selectbox(
        "Business type (for GL)",
        ["Software / SaaS", "Hardware / physical products", "Professional services"],
        index=0
    )

    capital_raised = st.number_input("Total capital raised (USD)", min_value=0.0, value=5_000_000.0, step=250_000.0, format="%.0f")

    st.subheader("EPL (derived from states)")
    primary_state = st.selectbox(
        "Primary state (HQ / majority employees)",
        ["CA","NY","IL","NJ","MA","WA","TX","FL","CO","GA","NC","PA","VA","AZ","Other"],
        index=0
    )
    pct_in_ca = st.slider("% of employees in California (optional)", min_value=0, max_value=100, value=0, step=5)
    pct_in_ca_val = float(pct_in_ca) if pct_in_ca > 0 else None

    st.subheader("Crime")
    regulated = st.toggle("Regulated industry (e.g., Fintech, HealthTech)", value=False)

try:
    cyber_df, dno_df, epl_uw_df, crime_df, gl_df, epl_reco_map = load_tables(WORKBOOK_PATH)
except FileNotFoundError:
    st.error(
        "Workbook not found. Place the Excel file next to this app and name it "
        f"'{WORKBOOK_PATH}', or update WORKBOOK_PATH in the code."
    )
    st.stop()

# Compute recommendations
gl_out = recommend_gl(gl_df, revenue, employees, biz_type)
cy_out = recommend_cyber(cyber_df, revenue)
do_out = recommend_dno(dno_df, capital_raised)
epl_out = recommend_epl(epl_uw_df, epl_reco_map, employees, primary_state, pct_in_ca_val)
cr_out = recommend_crime(crime_df, capital_raised, regulated)

# Layout
c1, c2 = st.columns(2, gap="large")

with c1:
    st.subheader("General Liability (GL)")
    st.json(gl_out, expanded=True)

    st.subheader("Cyber")
    st.json(cy_out, expanded=True)

    st.subheader("Crime")
    st.json(cr_out, expanded=True)

with c2:
    st.subheader("Directors & Officers (D&O)")
    st.json(do_out, expanded=True)

    st.subheader("Employment Practices Liability (EPL)")
    st.json(epl_out, expanded=True)

st.divider()
st.markdown(
    """
**Notes / guardrails**
- EPL “state group” is **derived** from the primary state (and CA footprint if provided) so users never need to pick an internal bucket.
- GL recommended limit uses the **GL GTM limit vs. UW limit** table (software vs hardware). For *Professional services*, the tool currently uses the software column (conservative default).
- If the workbook structure changes (especially the Exposures sheet row positions), re-anchor the EPL recommended table extraction.
"""
)
